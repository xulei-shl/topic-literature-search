"""万方导出结果处理测试。"""

import sys
import unittest
from pathlib import Path
from tempfile import TemporaryDirectory

import openpyxl

ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from tests.script_loader import load_script_module

SCRIPT_DIR = ROOT_DIR / "wanfang-search" / "scripts"
_cli_module = load_script_module(SCRIPT_DIR, "cli", "wanfang_cli_module")
_export_processor_module = load_script_module(
    SCRIPT_DIR,
    "export_processor",
    "wanfang_export_processor_module",
)

create_parser = _cli_module.create_parser
normalize_date_range = _cli_module.normalize_date_range
parse_cli_date = _cli_module.parse_cli_date
ExportResultProcessor = _export_processor_module.ExportResultProcessor


class ExportResultProcessorTestCase(unittest.TestCase):
    """验证万方导出结果本地处理逻辑。"""

    def setUp(self) -> None:
        self.processor = ExportResultProcessor()

    def test_parse_reference_txt_supports_multiline_items(self) -> None:
        """应按编号解析多行引文。"""
        with TemporaryDirectory() as temp_dir:
            txt_path = Path(temp_dir) / "refs.txt"
            txt_path.write_text(
                "[1]第一条引文\n延续内容\n[2]第二条引文\n[3]第三条引文",
                encoding="utf-8",
            )

            references = self.processor.parse_reference_txt(txt_path)

        self.assertEqual(references, ["[1]第一条引文 延续内容", "[2]第二条引文", "[3]第三条引文"])

    def test_sanitize_export_excel_removes_type_label_row_and_blank_column(self) -> None:
        """应删除类型标签首行与首列空白。"""
        with TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            raw_excel_path = temp_path / "raw.xls"
            cleaned_excel_path = temp_path / "cleaned.xlsx"
            raw_excel_path.write_text(
                """
                <table>
                    <tr><td>期刊</td><td></td><td></td></tr>
                    <tr><td></td><td>序号</td><td>题名</td><td>作者</td></tr>
                    <tr><td></td><td>1</td><td>文章一</td><td>张三</td></tr>
                    <tr><td></td><td>2</td><td>文章二</td><td>李四</td></tr>
                </table>
                """,
                encoding="utf-8",
            )

            result_path = self.processor.sanitize_export_excel(raw_excel_path, cleaned_excel_path)
            workbook = openpyxl.load_workbook(result_path)
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))

        self.assertEqual(rows[0], ("序号", "题名", "作者"))
        self.assertEqual(rows[1], ("1", "文章一", "张三"))
        self.assertEqual(rows[2], ("2", "文章二", "李四"))

    def test_enrich_batch_excel_adds_reference_column(self) -> None:
        """应为批次表格补齐参考格式列。"""
        with TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            cleaned_excel_path = temp_path / "batch001_cleaned.xlsx"
            txt_path = temp_path / "batch001.txt"
            output_path = temp_path / "batch001_enriched.xlsx"

            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(["序号", "题名", "作者"])
            sheet.append([1, "文章一", "张三"])
            sheet.append([2, "文章二", "李四"])
            workbook.save(cleaned_excel_path)

            txt_path.write_text("[1]引文一\n[2]引文二", encoding="utf-8")

            result_path = self.processor.enrich_batch_excel(cleaned_excel_path, txt_path, output_path)
            workbook = openpyxl.load_workbook(result_path)
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))

        self.assertEqual(rows[0], ("序号", "题名", "作者", "参考格式"))
        self.assertEqual(rows[1], (1, "文章一", "张三", "[1]引文一"))
        self.assertEqual(rows[2], (2, "文章二", "李四", "[2]引文二"))

    def test_merge_batch_excels_keeps_order(self) -> None:
        """应按批次顺序合并结果。"""
        with TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            first_path = temp_path / "first.xlsx"
            second_path = temp_path / "second.xlsx"
            final_path = temp_path / "merged.xlsx"

            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(["题名", "参考格式"])
            sheet.append(["文章一", "[1]引文一"])
            workbook.save(first_path)

            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(["题名", "参考格式"])
            sheet.append(["文章二", "[1]引文二"])
            workbook.save(second_path)

            result_path = self.processor.merge_batch_excels([first_path, second_path], final_path)
            workbook = openpyxl.load_workbook(result_path)
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))

        self.assertEqual(rows[1], ("文章一", "[1]引文一"))
        self.assertEqual(rows[2], ("文章二", "[1]引文二"))


class CliValidationTestCase(unittest.TestCase):
    """验证 CLI 参数校验。"""

    def test_parse_cli_date_accepts_year_only(self) -> None:
        self.assertEqual(parse_cli_date("2024", "start"), "2024")
        self.assertEqual(parse_cli_date("2025", "end"), "2025")

    def test_parse_cli_date_rejects_full_date(self) -> None:
        with self.assertRaises(ValueError):
            parse_cli_date("2024-03-12", "end")

    def test_normalize_date_range_rejects_reversed_range(self) -> None:
        with self.assertRaises(ValueError):
            normalize_date_range("2025", "2024")

    def test_parser_has_only_login_and_advanced_search(self) -> None:
        parser = create_parser()
        with self.assertRaises(SystemExit):
            parser.parse_args(["search", "新青年"])

    def test_advanced_search_max_download_defaults_to_none(self) -> None:
        parser = create_parser()
        args = parser.parse_args(["advanced-search", "--query", "新青年"])
        self.assertIsNone(args.max_download)


if __name__ == "__main__":
    unittest.main()
