"""CNKI 导出结果处理测试。"""

import sys
import unittest
from pathlib import Path
from tempfile import TemporaryDirectory

import openpyxl

SCRIPT_DIR = Path(__file__).resolve().parents[1] / "cnki-search" / "scripts"
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

from cli import create_parser, normalize_date_range, parse_cli_date
from export_processor import ExportResultProcessor


class ExportResultProcessorTestCase(unittest.TestCase):
    """验证导出结果本地处理逻辑。"""

    def setUp(self) -> None:
        """初始化测试对象。"""
        self.processor = ExportResultProcessor()

    def test_parse_reference_txt_supports_multiline_items(self) -> None:
        """应按编号解析多行引文。"""
        with TemporaryDirectory() as temp_dir:
            txt_path = Path(temp_dir) / "refs.txt"
            txt_path.write_text(
                "[1]第一条引文\n延续内容\n[2]第二条引文\n[3]第三条引文",
                encoding="utf-8",
            )

            references = self.processor.parse_reference_txt(txt_path)

        self.assertEqual(
            references,
            ["[1]第一条引文 延续内容", "[2]第二条引文", "[3]第三条引文"],
        )

    def test_enrich_batch_excel_adds_reference_column(self) -> None:
        """应为批次表格补齐参考格式列。"""
        with TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            raw_excel_path = temp_path / "batch001.xls"
            cleaned_excel_path = temp_path / "batch001_cleaned.xlsx"
            txt_path = temp_path / "batch001.txt"
            output_path = temp_path / "batch001_enriched.xlsx"

            raw_excel_path.write_text(
                """
                <table>
                    <tr><th>题名</th><th>作者</th></tr>
                    <tr><td>文章一</td><td>张三</td></tr>
                    <tr><td>文章二</td><td>李四</td></tr>
                </table>
                """,
                encoding="utf-8",
            )
            txt_path.write_text("[1]引文一\n[2]引文二", encoding="utf-8")

            self.processor.sanitize_export_excel(raw_excel_path, cleaned_excel_path)
            result_path = self.processor.enrich_batch_excel(cleaned_excel_path, txt_path, output_path)

            workbook = openpyxl.load_workbook(result_path)
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))

        self.assertEqual(rows[0], ("题名", "作者", "参考格式"))
        self.assertEqual(rows[1], ("文章一", "张三", "[1]引文一"))
        self.assertEqual(rows[2], ("文章二", "李四", "[2]引文二"))

    def test_merge_batch_excels_keeps_order(self) -> None:
        """应按批次顺序合并结果。"""
        with TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            first_path = temp_path / "first.xlsx"
            second_path = temp_path / "second.xlsx"
            final_path = temp_path / "merged.xlsx"

            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(["题名", "参考格式"])
            sheet.append(["文章一", "[1]引文一"])
            workbook.save(first_path)

            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(["题名", "参考格式"])
            sheet.append(["文章二", "[1]引文二"])
            workbook.save(second_path)

            result_path = self.processor.merge_batch_excels([first_path, second_path], final_path)

            workbook = openpyxl.load_workbook(result_path)
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))

        self.assertEqual(rows[1], ("文章一", "[1]引文一"))
        self.assertEqual(rows[2], ("文章二", "[1]引文二"))

    def test_sanitize_export_excel_supports_mixed_headers_and_keeps_order(self) -> None:
        """应支持混合表头并保留原始数据顺序。"""
        with TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            raw_excel_path = temp_path / "mixed.xls"
            cleaned_excel_path = temp_path / "mixed_cleaned.xlsx"

            raw_excel_path.write_text(
                """
                <table>
                    <tr><td>SrcDatabase-来源库</td><td>Author-作者</td><td>Title-题名</td><td>Keyword-关键词</td></tr>
                    <tr><td>期刊</td><td>作者甲</td><td>题名甲</td><td>关键词甲</td></tr>
                    <tr><td>SrcDatabase-来源库</td><td>Title-题名</td><td>Author-作者</td><td>Source-文献来源</td><td>DOI-DOI</td></tr>
                    <tr><td>报纸</td><td>题名乙</td><td>作者乙</td><td>来源乙</td><td>DOI乙</td></tr>
                </table>
                """,
                encoding="utf-8",
            )

            result_path = self.processor.sanitize_export_excel(raw_excel_path, cleaned_excel_path)

            workbook = openpyxl.load_workbook(result_path)
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))

        headers = rows[0]
        first_data = rows[1]
        second_data = rows[2]

        self.assertEqual(len(rows), 3)
        self.assertIn("SrcDatabase-来源库", headers)
        self.assertIn("Author-作者", headers)
        self.assertIn("Title-题名", headers)
        self.assertIn("Source-文献来源", headers)
        self.assertIn("DOI-DOI", headers)
        self.assertEqual(first_data[headers.index("Title-题名")], "题名甲")
        self.assertEqual(second_data[headers.index("Title-题名")], "题名乙")
        self.assertEqual(first_data[headers.index("Author-作者")], "作者甲")
        self.assertEqual(second_data[headers.index("Author-作者")], "作者乙")
        self.assertEqual(second_data[headers.index("Source-文献来源")], "来源乙")
        self.assertEqual(second_data[headers.index("DOI-DOI")], "DOI乙")


class CliValidationTestCase(unittest.TestCase):
    """验证 CLI 参数校验。"""

    def test_advanced_search_date_range_defaults_to_none(self) -> None:
        """未传日期范围时不应注入默认值。"""
        parser = create_parser()

        args = parser.parse_args(["advanced-search", "--query", "新青年"])

        self.assertIsNone(args.date_from)
        self.assertIsNone(args.date_to)

    def test_parse_cli_date_accepts_year_only(self) -> None:
        """高级检索日期应只接受年份。"""
        self.assertEqual(parse_cli_date("2024", "start"), "2024")
        self.assertEqual(parse_cli_date("2025", "end"), "2025")

    def test_parse_cli_date_rejects_month_or_full_date(self) -> None:
        """月份或完整日期输入应被拒绝。"""
        with self.assertRaises(ValueError):
            parse_cli_date("2024-2", "end")

        with self.assertRaises(ValueError):
            parse_cli_date("2024-03-12", "end")

    def test_parse_cli_date_rejects_invalid_date(self) -> None:
        """应拒绝非法日期。"""
        with self.assertRaises(ValueError):
            parse_cli_date("abcd", "end")

    def test_normalize_date_range_rejects_reversed_range(self) -> None:
        """起始日期晚于结束日期时应报错。"""
        with self.assertRaises(ValueError):
            normalize_date_range("2025", "2024")

    def test_parser_accepts_new_date_range_arguments(self) -> None:
        """高级检索应接受新的日期范围参数。"""
        parser = create_parser()

        args = parser.parse_args(["advanced-search", "--query", "新青年", "--date-from", "2024", "--date-to", "2025"])

        self.assertEqual(args.date_from, "2024")
        self.assertEqual(args.date_to, "2025")

    def test_parser_accepts_include_no_fulltext_flag(self) -> None:
        """高级检索应支持取消仅看有全文。"""
        parser = create_parser()

        args = parser.parse_args(["advanced-search", "--query", "新青年", "--include-no-fulltext"])

        self.assertTrue(args.include_no_fulltext)


if __name__ == "__main__":
    unittest.main()
