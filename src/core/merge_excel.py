"""Excel 合并工具 - 合并多个数据源的 Excel 文件并预处理。"""

import glob
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

import yaml
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent

SOURCE_NAMESPACES = {
    "cnki": "cnki-search",
    "wanfang": "wanfang-search",
    "vp": "vp-search",
}


def load_config() -> dict[str, Any]:
    """加载配置文件。"""
    config_path = PROJECT_ROOT / "config" / "config.yaml"
    with open(config_path, encoding="utf-8") as f:
        return yaml.safe_load(f)


def find_merged_file(namespace: str, query: str) -> Path | None:
    """查找指定数据源的 merged.xlsx 文件。

    Args:
        namespace: 数据源命名空间 (如 cnki-search)。
        query: 检索关键词。

    Returns:
        文件路径，如果不存在则返回 None。
    """
    output_dir = PROJECT_ROOT / "outputs" / namespace / query
    if not output_dir.exists():
        return None

    pattern = str(output_dir / "*-merged.xlsx")
    matches = glob.glob(pattern)
    if not matches:
        return None
    return Path(matches[0])


def convert_simplified(text: str) -> str:
    """将繁体中文转换为简体中文。

    Args:
        text: 待转换文本。

    Returns:
        转换后的文本。
    """
    try:
        import opencc
        converter = opencc.OpenCC("t2s")
        return converter.convert(text)
    except ImportError:
        return text


import re


REFERENCE_NUMBER_PATTERN = re.compile(r"^\[\d+\]\s*")


def clean_reference_number(text: str | None) -> str | None:
    """去掉参考格式开头的序号，如 [1]、[123] 等。

    Args:
        text: 参考格式文本。

    Returns:
        清理后的文本。
    """
    if text is None:
        return None
    return REFERENCE_NUMBER_PATTERN.sub("", text)


def check_contains_xinnianqing(text: str | None) -> str:
    """检查文本是否严格包含《新青年》。

    Args:
        text: 待检查文本。

    Returns:
        "有" 或 "无"。
    """
    if text is None:
        return "无"
    return "有" if "《新青年》" in str(text) else "无"


def normalize_header(header: str) -> str:
    """标准化表头：去除所有空格。

    Args:
        header: 原始表头。

    Returns:
        去除空格后的表头。
    """
    return header.replace(" ", "") if header else header


def read_excel_with_mapping(file_path: Path, source_mapping: dict[str, str]) -> list[dict]:
    """读取 Excel 文件并按照映射关系转换列名。

    Args:
        file_path: Excel 文件路径。
        source_mapping: 源字段到目标字段的映射。

    Returns:
        转换后的数据行列表。
    """
    wb = load_workbook(file_path, read_only=True)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    normalized_header_map = {normalize_header(h): idx for idx, h in enumerate(headers) if h}

    normalized_mapping = {normalize_header(k): v for k, v in source_mapping.items()}

    target_headers = list(load_config()["target_fields"])

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue

        converted_row = {}
        for source_col, target_col in normalized_mapping.items():
            if source_col in normalized_header_map:
                value = row[normalized_header_map[source_col]]
                if isinstance(value, str):
                    value = convert_simplified(value)
                    if target_col == "参考格式":
                        value = clean_reference_number(value)
                converted_row[target_col] = value

        converted_row["检索数据库"] = file_path.parent.parent.name.replace("-search", "")

        rows.append(converted_row)

    wb.close()
    return rows


def generate_merge_report(
    source_info: list[dict],
    output_path: Path,
    all_rows: list[dict],
    query: str,
) -> str:
    """生成合并结果的 Markdown 报告并保存。

    Args:
        source_info: 各数据源信息（target, file_path, row_count）。
        output_path: 合并后的 Excel 文件路径。
        all_rows: 所有数据行。
        query: 检索关键词。

    Returns:
        报告文件路径。
    """
    from collections import Counter

    title_cnt = Counter()
    abstract_cnt = Counter()
    keyword_cnt = Counter()
    combos = Counter()

    for row in all_rows:
        t = check_contains_xinnianqing(row.get("题名"))
        a = check_contains_xinnianqing(row.get("摘要"))
        k = check_contains_xinnianqing(row.get("关键词"))
        title_cnt[t] += 1
        abstract_cnt[a] += 1
        keyword_cnt[k] += 1
        combos[(t, a, k)] += 1

    total = len(all_rows)

    lines = [
        "# 合并结果报告",
        "",
        f"- **检索词**: {query}",
        f"- **生成时间**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"- **总记录数**: {total}",
        "",
        "---",
        "",
        "## 数据源",
        "",
        "| 数据库 | 原始文件 | 记录数 |",
        "|--------|----------|--------|",
    ]
    for info in source_info:
        lines.append(f"| {info['target']} | {info['file_path']} | {info['row_count']} |")

    lines += [
        "",
        "## 合并输出",
        "",
        f"- **输出文件**: {output_path}",
        "",
        "---",
        "",
        "## 各列《新青年》包含情况",
        "",
        "| 列 | 有 | 无 |",
        "|-----|----|----|",
        f"| 题名含《新青年》 | {title_cnt.get('有', 0)} | {title_cnt.get('无', 0)} |",
        f"| 摘要含《新青年》 | {abstract_cnt.get('有', 0)} | {abstract_cnt.get('无', 0)} |",
        f"| 关键词含《新青年》 | {keyword_cnt.get('有', 0)} | {keyword_cnt.get('无', 0)} |",
        "",
        "## 三列组配统计",
        "",
        "| 题名含《新青年》 | 摘要含《新青年》 | 关键词含《新青年》 | 数量 | 占比 |",
        "|-----------------|-----------------|-----------------|------|------|",
    ]
    for t in ("有", "无"):
        for a in ("有", "无"):
            for k in ("有", "无"):
                count = combos.get((t, a, k), 0)
                pct = f"{count / total * 100:.1f}%" if total > 0 else "0%"
                lines.append(f"| {t} | {a} | {k} | {count} | {pct} |")

    report_path = output_path.with_suffix(".md")
    report_path.write_text("\n".join(lines), encoding="utf-8")
    return str(report_path)


def merge_excel_files(query: str, targets: list[str]) -> str:
    """合并多个数据源的 Excel 文件。

    Args:
        query: 检索关键词。
        targets: 需要合并的数据源列表。

    Returns:
        合并后的文件路径。
    """
    config = load_config()
    source_mappings = config["source_mappings"]
    target_fields = config["target_fields"]

    all_rows = []
    source_info = []

    for target in targets:
        namespace = SOURCE_NAMESPACES.get(target)
        if not namespace:
            print(f"警告: 未知的数据源 {target}")
            continue

        merged_file = find_merged_file(namespace, query)
        if not merged_file:
            print(f"警告: 找不到 {namespace}/{query} 的 merged.xlsx，跳过")
            continue

        mapping = source_mappings.get(target, {})
        print(f"读取: {merged_file}")
        rows = read_excel_with_mapping(merged_file, mapping)
        print(f"  -> 读取 {len(rows)} 行")
        source_info.append({"target": target, "file_path": str(merged_file), "row_count": len(rows)})
        all_rows.extend(rows)

    if not all_rows:
        raise ValueError("没有找到任何可合并的数据")

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active

    for idx, field in enumerate(target_fields, start=1):
        ws.cell(row=1, column=idx, value=field)

    ws.cell(row=1, column=len(target_fields) + 1, value="题名含有《新青年》")
    ws.cell(row=1, column=len(target_fields) + 2, value="摘要含有《新青年》")
    ws.cell(row=1, column=len(target_fields) + 3, value="关键词含有《新青年》")

    for row_idx, row_data in enumerate(all_rows, start=2):
        for col_idx, field in enumerate(target_fields, start=1):
            value = row_data.get(field, "")
            ws.cell(row=row_idx, column=col_idx, value=value)

        title_val = row_data.get("题名", "")
        abstract_val = row_data.get("摘要", "")
        keyword_val = row_data.get("关键词", "")

        ws.cell(row=row_idx, column=len(target_fields) + 1, value=check_contains_xinnianqing(title_val))
        ws.cell(row=row_idx, column=len(target_fields) + 2, value=check_contains_xinnianqing(abstract_val))
        ws.cell(row=row_idx, column=len(target_fields) + 3, value=check_contains_xinnianqing(keyword_val))

    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    query_slug = "".join(c for c in query if c.isalnum() or c in (" ", "-", "_")).strip()[:30]
    output_filename = f"{timestamp}-{query_slug}-combined.xlsx"
    output_dir = PROJECT_ROOT / "outputs"
    output_path = output_dir / output_filename

    wb.save(output_path)
    print(f"合并完成: {output_path} (共 {len(all_rows)} 行)")

    report_path = generate_merge_report(source_info, output_path, all_rows, query)
    print(f"报告生成: {report_path}")

    return str(output_path)


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("用法: python -m src.core.merge_excel <query> <targets...>")
        sys.exit(1)

    query = sys.argv[1]
    targets = sys.argv[2:]
    merge_excel_files(query, targets)