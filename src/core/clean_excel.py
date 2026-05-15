"""Excel 去重工具 - 基于题名去重，跨库优先保留更完整记录。"""

import glob
import re
from collections import Counter
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook, Workbook

PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent


@dataclass
class _RowInfo:
    data: list
    source_db: str
    non_null_count: int


def _count_non_null(row: list) -> int:
    return sum(1 for v in row if v is not None and str(v).strip() != "")


def _write_sheet(ws, headers: list, rows: list):
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)


def generate_clean_report(
    input_path: Path,
    dedup_path: Path,
    filtered_path: Path,
    headers: list,
    original_count: int,
    count_external: int,
    count_empty_title: int,
    count_duplicate: int,
    kept_rows: list,
    query: str,
) -> str:
    """生成去重结果的 Markdown 报告并保存。

    Args:
        input_path: 输入文件路径。
        dedup_path: 去重后文件路径。
        filtered_path: 过滤数据文件路径。
        headers: 表头列表。
        original_count: 原始总行数（含被过滤行）。
        count_external: 外文过滤行数。
        count_empty_title: 空题名过滤行数。
        count_duplicate: 跨库重复淘汰行数。
        kept_rows: 保留的数据行。
        query: 检索关键词。

    Returns:
        报告文件路径。
    """
    kept_count = len(kept_rows)

    try:
        title_col = headers.index("题名含有《新青年》")
    except ValueError:
        title_col = -1
    try:
        abstract_col = headers.index("摘要含有《新青年》")
    except ValueError:
        abstract_col = -1
    try:
        keyword_col = headers.index("关键词含有《新青年》")
    except ValueError:
        keyword_col = -1

    has_xnq_columns = all(c >= 0 for c in (title_col, abstract_col, keyword_col))

    title_cnt: Counter[str] = Counter()
    abstract_cnt: Counter[str] = Counter()
    keyword_cnt: Counter[str] = Counter()
    combos: Counter[tuple[str, str, str]] = Counter()

    for row in kept_rows:
        t = str(row[title_col]) if has_xnq_columns and title_col < len(row) else "无"
        a = str(row[abstract_col]) if has_xnq_columns and abstract_col < len(row) else "无"
        k = str(row[keyword_col]) if has_xnq_columns and keyword_col < len(row) else "无"
        title_cnt[t] += 1
        abstract_cnt[a] += 1
        keyword_cnt[k] += 1
        combos[(t, a, k)] += 1

    total_filtered = count_external + count_empty_title + count_duplicate
    lines = [
        "# 去重结果报告",
        "",
        f"- **检索词**: {query}",
        f"- **生成时间**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        "",
        "---",
        "",
        "## 去重概览",
        "",
        "| 指标 | 数值 |",
        "|------|------|",
        f"| 输入文件 | {input_path} |",
        f"| 输出结果文件 | {dedup_path} |",
        f"| 输出过滤文件 | {filtered_path} |",
        f"| 原始总行数 | {original_count} |",
        f"| 过滤（外文来源） | {count_external} |",
        f"| 过滤（题名为空） | {count_empty_title} |",
        f"| 过滤（跨库重复淘汰） | {count_duplicate} |",
        f"| 合计划除行数 | {total_filtered} |",
        f"| 保留行数 | {kept_count} |",
        "",
        "---",
    ]

    if has_xnq_columns:
        lines += [
            "## 各列《新青年》包含情况（去重后）",
            "",
            "| 列 | 有 | 无 |",
            "|-----|----|----|",
            f"| 题名含《新青年》 | {title_cnt.get('有', 0)} | {title_cnt.get('无', 0)} |",
            f"| 摘要含《新青年》 | {abstract_cnt.get('有', 0)} | {abstract_cnt.get('无', 0)} |",
            f"| 关键词含《新青年》 | {keyword_cnt.get('有', 0)} | {keyword_cnt.get('无', 0)} |",
            "",
            "## 三列组配统计（去重后）",
            "",
            "| 题名含《新青年》 | 摘要含《新青年》 | 关键词含《新青年》 | 数量 | 占比 |",
            "|-----------------|-----------------|-----------------|------|------|",
        ]
        for t in ("有", "无"):
            for a in ("有", "无"):
                for k in ("有", "无"):
                    count = combos.get((t, a, k), 0)
                    pct = f"{count / kept_count * 100:.1f}%" if kept_count > 0 else "0%"
                    lines.append(f"| {t} | {a} | {k} | {count} | {pct} |")
    else:
        lines.append("*（合并文件中未包含《新青年》三列统计，跳过）*")

    report_path = dedup_path.with_suffix(".md")
    report_path.write_text("\n".join(lines), encoding="utf-8")
    return str(report_path)


def normalize_title(title: str) -> str:
    """标准化题名用于去重匹配。

    处理全角/半角标点、空格差异，确保不同来源的相似题名能正确匹配。
    """
    if not title:
        return ""
    ftoc = str.maketrans(
        '，。：；？！""\'\'（）【】《》、·——…',
        ',.:;?!""\'\'()[]<>,.--.'
    )
    text = title.translate(ftoc)
    text = re.sub(r'\s+', '', text)
    return text.lower()


def normalize_query_slug(query: str) -> str:
    """将查询词转换为文件名中使用的 slug（与 merge_excel.py 保持一致）。"""
    return "".join(c for c in query if c.isalnum() or c in (" ", "-", "_")).strip()[:30]


def find_latest_combined_file(query: str) -> Optional[Path]:
    """根据查询词查找最新的合并 Excel 文件。"""
    slug = normalize_query_slug(query)
    pattern = str(PROJECT_ROOT / "outputs" / f"*-{slug}-combined.xlsx")
    matches = glob.glob(pattern)
    if not matches:
        return None

    def extract_timestamp(path: Path) -> str:
        name = path.name
        parts = name.split("-")
        if len(parts) >= 2:
            return parts[0] + parts[1]
        return ""

    matches.sort(key=lambda p: extract_timestamp(Path(p)), reverse=True)
    return Path(matches[0])


def deduplicate_by_title(
    input_path: Path, output_path: Optional[Path] = None, query: str = ""
) -> Path:
    """根据题名去重 Excel 文件，同库全保留，跨库择优保留。

    Returns:
        输出去重后文件路径（过滤文件自动生成在同目录下）。
    """
    if output_path is None:
        stem = input_path.stem
        output_path = input_path.parent / f"{stem}_deduplicated.xlsx"
    filtered_path = input_path.parent / f"{input_path.stem}_filtered.xlsx"

    wb_in = load_workbook(input_path, read_only=True)
    ws_in = wb_in.active

    headers = [cell.value for cell in ws_in[1]]
    if "题名" not in headers:
        raise ValueError("输入的 Excel 中缺少 '题名' 列，无法去重")

    title_col_idx = headers.index("题名")
    source_col_idx = headers.index("来源库") if "来源库" in headers else None

    filtered_external = []
    filtered_empty_title = []
    title_groups: dict[str, list[_RowInfo]] = {}

    original_count = 0
    count_external = 0

    for row in ws_in.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        original_count += 1
        row_list = list(row)

        # 外文过滤
        if (
            source_col_idx is not None
            and row_list[source_col_idx] is not None
            and "外文" in str(row_list[source_col_idx])
        ):
            count_external += 1
            filtered_external.append(row_list)
            continue

        # 空题名
        if row_list[title_col_idx] is None:
            filtered_empty_title.append(row_list)
            continue

        # 正常行：按标准化题名分组
        title = str(row_list[title_col_idx])
        normalized = normalize_title(title)
        source_db = (
            str(row_list[source_col_idx])
            if source_col_idx is not None and row_list[source_col_idx] is not None
            else ""
        )
        non_null_count = _count_non_null(row_list)
        title_groups.setdefault(normalized, []).append(
            _RowInfo(data=row_list, source_db=source_db, non_null_count=non_null_count)
        )

    wb_in.close()

    # 跨库去重：同库全保留，跨库择优
    deduped_rows = []
    filtered_duplicates = []

    for group in title_groups.values():
        db_groups: dict[str, list[_RowInfo]] = {}
        for info in group:
            db_groups.setdefault(info.source_db, []).append(info)

        if len(db_groups) == 1:
            for info in group:
                deduped_rows.append(info.data)
        else:
            best = max(group, key=lambda x: x.non_null_count)
            for info in group:
                if info is best:
                    deduped_rows.append(info.data)
                else:
                    filtered_duplicates.append(info.data)

    count_empty_title = len(filtered_empty_title)
    count_duplicate = len(filtered_duplicates)

    # 写出去重文件
    wb_dedup = Workbook()
    _write_sheet(wb_dedup.active, headers, deduped_rows)
    wb_dedup.save(output_path)
    print(
        f"去重完成: {output_path} "
        f"(原始 {original_count} 行 -> 保留 {len(deduped_rows)} 行)"
    )

    # 写入过滤文件
    wb_filtered = Workbook()
    sheet_specs = [
        ("外文", filtered_external),
        ("空题名", filtered_empty_title),
        ("重复", filtered_duplicates),
    ]
    first = True
    for name, rows in sheet_specs:
        if rows:
            if first:
                ws = wb_filtered.active
                ws.title = name
                first = False
            else:
                ws = wb_filtered.create_sheet(title=name)
            _write_sheet(ws, headers, rows)
    if first:
        _write_sheet(wb_filtered.active, headers, [])
        wb_filtered.active.title = "过滤数据"
    wb_filtered.save(filtered_path)
    print(
        f"过滤数据: {filtered_path} "
        f"(外文 {len(filtered_external)} 行, "
        f"空题名 {len(filtered_empty_title)} 行, "
        f"重复 {len(filtered_duplicates)} 行)"
    )

    if query:
        report_path = generate_clean_report(
            input_path=input_path,
            dedup_path=output_path,
            filtered_path=filtered_path,
            headers=headers,
            original_count=original_count,
            count_external=count_external,
            count_empty_title=count_empty_title,
            count_duplicate=count_duplicate,
            kept_rows=deduped_rows,
            query=query,
        )
        print(f"报告生成: {report_path}")

    return output_path


def clean_excel(query: str, input_path: Optional[Path] = None) -> str:
    """对第一步合并的结果进行去重。

    Args:
        query: 检索关键词，用于自动查找合并文件。
        input_path: 可选的输入文件路径，若提供则直接使用。

    Returns:
        输出去重后的文件路径。
    """
    if input_path is None:
        combined_file = find_latest_combined_file(query)
        if combined_file is None:
            raise FileNotFoundError(f"未找到查询词 '{query}' 对应的合并文件，请先运行 merge-excel")
        input_path = combined_file
    else:
        input_path = Path(input_path)
        if not input_path.exists():
            raise FileNotFoundError(f"输入文件不存在: {input_path}")

    return str(deduplicate_by_title(input_path, query=query))


if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("用法: python -m src.core.clean_excel <query> [input_file]")
        sys.exit(1)
    query = sys.argv[1]
    input_file = Path(sys.argv[2]) if len(sys.argv) > 2 else None
    clean_excel(query, input_file)
