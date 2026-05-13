"""LLM 过滤评估模块 - 对去重后的 Excel 进行主题相关性评估。"""

import asyncio
import json
import random
import re
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import aiohttp
import yaml
from openpyxl import load_workbook, Workbook

PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent

LLM_COLUMNS = [
    "is_target_magazine",
    "relevance_score",
    "relevance_level",
    "reasoning",
    "historical_keywords",
    "LLM返回JSON",
    "LLM错误信息",
]

RETRY_DELAYS = [3, 6, 10]
MAX_RETRIES_PER_CONFIG = 3
DEFAULT_BATCH_SIZE = 5
DEFAULT_MAX_ROUNDS = 3


@dataclass
class LLMConfig:
    """LLM 配置组（key + base_url + model）。"""

    api_key: str
    base_url: str
    model: str


@dataclass
class LLMResult:
    """单条数据的 LLM 评估结果。"""

    is_target_magazine: bool | None = None
    relevance_score: int | None = None
    relevance_level: str | None = None
    reasoning: str | None = None
    historical_keywords: list | None = None
    raw_json: str = ""
    error: str = ""

    def is_success(self) -> bool:
        return self.error == "" and self.raw_json != ""


def load_config() -> dict[str, Any]:
    """加载配置文件。"""
    config_path = PROJECT_ROOT / "config" / "config.yaml"
    with open(config_path, encoding="utf-8") as f:
        return yaml.safe_load(f)


def load_env_config() -> dict[str, str]:
    """从 .env 加载 LLM 相关配置。"""
    env_path = PROJECT_ROOT / ".env"
    config = {}
    if env_path.exists():
        with open(env_path, encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line and "=" in line and not line.startswith("#"):
                    key, value = line.split("=", 1)
                    config[key.strip()] = value.strip()
    return config


def parse_llm_configs(env_config: dict[str, str]) -> list[LLMConfig]:
    """解析 LLM_CONFIG 配置。

    格式：key|base_url|model,key|base_url|model
    """
    config_str = env_config.get("LLM_CONFIG", "")
    if not config_str:
        raise ValueError("LLM_CONFIG 未在 .env 中配置")

    configs = []
    for group in config_str.split(","):
        parts = group.strip().split("|")
        if len(parts) != 3:
            raise ValueError(f"LLM_CONFIG 格式错误: {group}，应为 key|base_url|model")
        configs.append(LLMConfig(api_key=parts[0].strip(), base_url=parts[1].strip(), model=parts[2].strip()))

    if not configs:
        raise ValueError("LLM_CONFIG 解析后为空")
    return configs


def get_batch_size(env_config: dict[str, str]) -> int:
    """获取批次大小。"""
    return int(env_config.get("LLM_BATCH_SIZE", str(DEFAULT_BATCH_SIZE)))


def get_max_rounds(env_config: dict[str, str]) -> int:
    """获取最大轮次。"""
    return int(env_config.get("LLM_MAX_ROUNDS", str(DEFAULT_MAX_ROUNDS)))


def load_prompt_template() -> str:
    """加载主题相关性评估的 Prompt 模板。"""
    prompt_path = PROJECT_ROOT / "config" / "主题相关性评估.md"
    with open(prompt_path, encoding="utf-8") as f:
        return f.read()


def build_prompt(template: str, title: str, keywords: str, abstract: str) -> list[dict[str, str]]:
    """构建发送给 LLM 的消息。

    Args:
        template: 主题相关性评估.md 的内容
        title: 题名
        keywords: 关键词
        abstract: 摘要

    Returns:
        messages 列表
    """
    content_parts = [
        template,
        "\n\n--- 文献信息 ---",
        f"\n【题名】{title}" if title else "\n【题名】无",
        f"\n【关键词】{keywords}" if keywords else "\n【关键词】无",
        f"\n【摘要】{abstract}" if abstract else "\n【摘要】无",
    ]
    return [{"role": "user", "content": "".join(content_parts)}]


async def call_llm_once(
    session: aiohttp.ClientSession,
    config: LLMConfig,
    messages: list[dict[str, str]],
    retry_count: int,
) -> LLMResult:
    """单次调用 LLM API。

    Args:
        session: aiohttp 会话
        config: LLM 配置
        messages: 消息列表
        retry_count: 当前重试次数（用于计算延迟）

    Returns:
        LLMResult
    """
    delay = RETRY_DELAYS[retry_count] if retry_count < len(RETRY_DELAYS) else RETRY_DELAYS[-1]
    if retry_count > 0:
        await asyncio.sleep(delay)

    url = f"{config.base_url.rstrip('/')}/v1/chat/completions"
    headers = {"Authorization": f"Bearer {config.api_key}", "Content-Type": "application/json"}
    payload = {"model": config.model, "messages": messages, "stream": False}

    try:
        async with session.post(url, json=payload, headers=headers, timeout=aiohttp.ClientTimeout(total=60)) as resp:
            if resp.status != 200:
                text = await resp.text()
                return LLMResult(error=f"HTTP {resp.status}: {text[:200]}")

            data = await resp.json()
            content = data.get("choices", [{}])[0].get("message", {}).get("content", "")
            return LLMResult(raw_json=content)
    except asyncio.TimeoutError:
        return LLMResult(error="请求超时")
    except Exception as e:
        return LLMResult(error=str(e)[:200])


async def call_llm_with_retry(
    session: aiohttp.ClientSession,
    configs: list[LLMConfig],
    messages: list[dict[str, str]],
) -> LLMResult:
    """使用多组配置重试调用 LLM。

    Args:
        session: aiohttp 会话
        configs: LLM 配置列表
        messages: 消息列表

    Returns:
        LLMResult（成功或失败）
    """
    shuffled_configs = configs.copy()
    random.shuffle(shuffled_configs)

    for config_group_idx in range(len(shuffled_configs)):
        config = shuffled_configs[config_group_idx]
        for retry in range(MAX_RETRIES_PER_CONFIG):
            result = await call_llm_once(session, config, messages, retry)
            if result.is_success():
                return result
            if retry < MAX_RETRIES_PER_CONFIG - 1:
                await asyncio.sleep(RETRY_DELAYS[retry])

        if config_group_idx < len(shuffled_configs) - 1:
            await asyncio.sleep(2)

    return result if 'result' in locals() else LLMResult(error="所有配置重试失败")


def parse_llm_response(result: LLMResult) -> LLMResult:
    """解析 LLM 返回的 JSON 响应。

    Args:
        result: 包含 raw_json 的 LLMResult

    Returns:
        解析后的 LLMResult（包含字段值或错误信息）
    """
    if not result.raw_json:
        return result

    try:
        cleaned = result.raw_json.strip()
        cleaned = re.sub(r"^```json\s*", "", cleaned)
        cleaned = re.sub(r"\s*```$", "", cleaned)
        cleaned = cleaned.strip()
        if not cleaned.startswith("{"):
            match = re.search(r"\{.*\}", cleaned, re.DOTALL)
            if match:
                cleaned = match.group(0)

        data = json.loads(cleaned)

        result.is_target_magazine = bool(data.get("is_target_magazine"))
        score = data.get("relevance_score")
        result.relevance_score = int(score) if isinstance(score, (int, str)) and str(score).isdigit() else None
        result.relevance_level = data.get("relevance_level", "")
        result.reasoning = (data.get("reasoning") or "")[:50]
        result.historical_keywords = data.get("historical_keywords", [])

        return result

    except json.JSONDecodeError as e:
        result.error = f"JSON 解析失败: {str(e)[:100]}"
        return result


def normalize_query_slug(query: str) -> str:
    """将查询词转换为文件名中使用的 slug。"""
    return "".join(c for c in query if c.isalnum() or c in (" ", "-", "_")).strip()[:30]


def find_latest_deduplicated_file(query: str) -> Path | None:
    """查找最新的去重 Excel 文件。"""
    slug = normalize_query_slug(query)
    pattern = str(PROJECT_ROOT / "outputs" / f"*-{slug}-*_deduplicated.xlsx")
    import glob as glob_module

    matches = glob_module.glob(pattern)
    if not matches:
        return None

    def extract_timestamp(path: Path) -> str:
        name = path.name
        parts = name.split("-")
        if len(parts) >= 2:
            return parts[0] + parts[1]
        return ""

    matches.sort(key=lambda p: extract_timestamp(Path(p)), reverse=True)
    return Path(matches[0])


async def process_single_row(
    session: aiohttp.ClientSession,
    configs: list[LLMConfig],
    prompt_template: str,
    row_data: dict[str, Any],
) -> LLMResult:
    """处理单行数据。

    Args:
        session: aiohttp 会话
        configs: LLM 配置列表
        prompt_template: Prompt 模板
        row_data: 行数据（包含题名、关键词、摘要）

    Returns:
        LLMResult
    """
    title = row_data.get("题名", "") or ""
    keywords = row_data.get("关键词", "") or ""
    abstract = row_data.get("摘要", "") or ""

    messages = build_prompt(prompt_template, title, keywords, abstract)
    result = await call_llm_with_retry(session, configs, messages)
    return parse_llm_response(result)


async def process_batch(
    session: aiohttp.ClientSession,
    configs: list[LLMConfig],
    prompt_template: str,
    rows: list[tuple[int, dict[str, Any]]],
    progress_callback=None,
) -> list[tuple[int, LLMResult]]:
    """异步并发处理一批数据。

    Args:
        session: aiohttp 会话
        configs: LLM 配置列表
        prompt_template: Prompt 模板
        rows: [(行索引, 行数据), ...]
        progress_callback: 进度回调函数

    Returns:
        [(行索引, LLMResult), ...]
    """
    async def process_one(row_idx: int, row_data: dict[str, Any]) -> tuple[int, LLMResult]:
        result = await process_single_row(session, configs, prompt_template, row_data)
        if progress_callback:
            progress_callback(1)
        return row_idx, result

    tasks = [process_one(idx, data) for idx, data in rows]
    results = await asyncio.gather(*tasks, return_exceptions=True)

    processed_results = []
    for i, result in enumerate(results):
        row_idx = rows[i][0]
        if isinstance(result, Exception):
            processed_results.append((row_idx, LLMResult(error=f"任务异常: {str(result)[:100]}")))
        else:
            processed_results.append(result)

    return processed_results


def read_excel_data(input_path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    """读取 Excel 数据。

    Returns:
        (表头列表, 行数据列表)
    """
    wb = load_workbook(input_path, read_only=True)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        row_dict = {headers[i]: row[i] for i in range(len(headers))}
        rows.append(row_dict)

    wb.close()
    return headers, rows


def write_excel_with_results(
    input_path: Path,
    headers: list[str],
    rows: list[dict[str, Any]],
    results: dict[int, LLMResult],
    output_path: Path,
) -> None:
    """将 LLM 结果写入 Excel。

    Args:
        input_path: 输入文件路径
        headers: 表头列表
        rows: 行数据列表
        results: {行索引: LLMResult}
        output_path: 输出文件路径
    """
    wb = Workbook()
    ws = wb.active

    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    for col_name in LLM_COLUMNS:
        ws.cell(row=1, column=len(headers) + LLM_COLUMNS.index(col_name) + 1, value=col_name)

    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))

        result = results.get(row_idx - 2)
        if result:
            ws.cell(row=row_idx, column=len(headers) + 1, value=str(result.is_target_magazine) if result.is_target_magazine is not None else "")
            ws.cell(row=row_idx, column=len(headers) + 2, value=result.relevance_score if result.relevance_score is not None else "")
            ws.cell(row=row_idx, column=len(headers) + 3, value=result.relevance_level or "")
            ws.cell(row=row_idx, column=len(headers) + 4, value=result.reasoning or "")
            ws.cell(row=row_idx, column=len(headers) + 5, value=json.dumps(result.historical_keywords, ensure_ascii=False) if result.historical_keywords else "[]")
            ws.cell(row=row_idx, column=len(headers) + 6, value=result.raw_json or "")
            ws.cell(row=row_idx, column=len(headers) + 7, value=result.error or "")

    wb.save(output_path)


def llm_filter(query: str, input_path: Path | None = None, batch_size: int | None = None) -> str:
    """对去重后的 Excel 进行 LLM 主题相关性评估。

    Args:
        query: 检索关键词
        input_path: 可选的输入文件路径
        batch_size: 可选的批次大小

    Returns:
        输出文件路径
    """
    env_config = load_env_config()
    configs = parse_llm_configs(env_config)
    prompt_template = load_prompt_template()

    if batch_size is None:
        batch_size = get_batch_size(env_config)
    max_rounds = get_max_rounds(env_config)

    if input_path is None:
        input_path = find_latest_deduplicated_file(query)
        if input_path is None:
            raise FileNotFoundError(f"未找到查询词 '{query}' 对应的去重文件，请先运行 clean-excel")
    else:
        input_path = Path(input_path)
        if not input_path.exists():
            raise FileNotFoundError(f"输入文件不存在: {input_path}")

    print(f"读取文件: {input_path}")
    headers, rows = read_excel_data(input_path)
    total_rows = len(rows)
    print(f"共 {total_rows} 行数据待处理")

    stem = input_path.stem
    output_path = input_path.parent / f"{stem}_llm_filtered.xlsx"

    results: dict[int, LLMResult] = {}
    error_indices = set(range(total_rows))
    submitted_count = 0
    completed_count = 0

    def save_progress():
        """保存当前进度到 Excel。"""
        write_excel_with_results(input_path, headers, rows, results, output_path)
        print(f"  [进度已保存] {output_path}")

    for round_num in range(1, max_rounds + 1):
        if not error_indices:
            break

        print(f"\n[第{round_num}轮/共{max_rounds}轮] 开始处理 {len(error_indices)} 个错误行...")

        current_batch = []
        round_error_indices = set()

        async def run_async():
            nonlocal current_batch, submitted_count, completed_count, round_error_indices

            connector = aiohttp.TCPConnector(limit=batch_size)
            async with aiohttp.ClientSession(connector=connector) as session:
                for idx in error_indices:
                    current_batch.append((idx, rows[idx]))
                    submitted_count += 1

                    if len(current_batch) >= batch_size or idx == list(error_indices)[-1]:
                        batch_to_process = current_batch
                        current_batch = []

                        batch_results = await process_batch(
                            session, configs, prompt_template, batch_to_process
                        )

                        for row_idx, result in batch_results:
                            results[row_idx] = result
                            completed_count += 1
                            if not result.is_success():
                                round_error_indices.add(row_idx)

                        print(f"[第{round_num}轮/共{max_rounds}轮] 已完成: {completed_count}/{total_rows}")
                        save_progress()

                if current_batch:
                    batch_results = await process_batch(
                        session, configs, prompt_template, current_batch
                    )
                    for row_idx, result in batch_results:
                        results[row_idx] = result
                        completed_count += 1
                        if not result.is_success():
                            round_error_indices.add(row_idx)
                    save_progress()

        asyncio.run(run_async())

        success_count = sum(1 for idx in error_indices if results.get(idx) and results[idx].is_success())
        error_count = len(error_indices) - success_count
        print(f"[第{round_num}轮/共{max_rounds}轮] 完成: 成功 {success_count}, 失败 {error_count}")

        error_indices = round_error_indices

    final_success = sum(1 for r in results.values() if r.is_success())
    final_error = sum(1 for r in results.values() if not r.is_success())

    print(f"\n处理完成: 总计 {total_rows} 行, 成功 {final_success}, 失败 {final_error}")
    print(f"输出文件: {output_path}")

    return str(output_path)


if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print("用法: python -m src.core.llm_filter <query> [input_file]")
        sys.exit(1)

    query = sys.argv[1]
    input_file = Path(sys.argv[2]) if len(sys.argv) > 2 else None
    llm_filter(query, input_file)